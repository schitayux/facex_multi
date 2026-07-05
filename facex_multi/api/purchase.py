"""
facex_multi.api.purchase
------------------------
Módulo de compras (Purchase Invoice) para FacEx Multi.
update_stock=1 siempre. Bodega y cuenta se resuelven automáticamente
según item_group + compañía o desde Item Default.
"""
from __future__ import annotations
import frappe
import json
from frappe.utils import today, getdate, flt


# ---------------------------------------------------------------------------
# Helpers de resolución
# ---------------------------------------------------------------------------

def _get_abbr(company: str) -> str:
    return frappe.db.get_value("Company", company, "abbr") or ""


def _resolve_item_warehouse(item_code: str, company: str) -> str:
    """Bodega para recepción: Item Default → patrón por grupo de ítem."""
    wh = frappe.db.get_value(
        "Item Default",
        {"parent": item_code, "company": company},
        "default_warehouse",
    )
    if wh:
        return wh
    abbr       = _get_abbr(company)
    item_group = frappe.db.get_value("Item", item_code, "item_group") or ""
    base       = "EL MOSQUETE" if abbr == "EMS" else "EL PANZER"
    idx        = {"ARMAS": "I", "MUNICIÓN": "II", "ACCESORIOS": "III"}.get(item_group)
    return f"{base} {idx} - {abbr}" if idx else ""


def _list_company_tax_templates(company: str) -> list:
    """Plantillas de impuestos de compra realmente activas y ligadas a la compañía."""
    rows = frappe.get_all(
        "Purchase Taxes and Charges Template",
        filters={"company": company, "disabled": 0},
        fields=["name", "is_default"],
        order_by="is_default desc, name asc",
    )
    return rows


def _resolve_purchase_tax_template(company: str, template_name: str = None) -> str:
    """Resuelve la plantilla de impuestos de compra a usar.

    Si template_name es una plantilla activa y ligada a la compañía, se respeta
    tal cual (selección explícita del usuario desde la lista real). Si no, se
    usa la marcada is_default=1, o la primera activa de la compañía.
    """
    if template_name and frappe.db.exists(
        "Purchase Taxes and Charges Template",
        {"name": template_name, "company": company, "disabled": 0},
    ):
        return template_name

    templates = _list_company_tax_templates(company)
    return templates[0].name if templates else ""


def _get_tax_rows(template_name: str) -> list:
    """Filas reales de la plantilla de impuestos, listas para doc.append('taxes', ...)."""
    if not template_name:
        return []
    from erpnext.controllers.accounts_controller import get_taxes_and_charges
    return get_taxes_and_charges("Purchase Taxes and Charges Template", template_name) or []


def _get_purchase_tax_rate(template_name: str) -> float:
    """Suma de porcentajes de impuesto de la plantilla indicada."""
    if not template_name:
        return 0.0
    rows = frappe.get_all(
        "Purchase Taxes and Charges",
        filters={"parent": template_name},
        fields=["rate"],
    )
    return sum(flt(r.rate) for r in rows)


def _get_naming_series_pi(company: str) -> list:
    abbr = _get_abbr(company)
    raw  = (
        frappe.db.get_value(
            "Property Setter",
            {"doc_type": "Purchase Invoice", "field_name": "naming_series", "property": "options"},
            "value",
        )
        or ""
    )
    all_s    = [s.strip() for s in raw.split("\n") if s.strip()]
    filtered = [s for s in all_s if abbr in s]
    return filtered or all_s


def _get_buying_price_list() -> str:
    """Lista de precios de compra configurada en Buying Settings (fallback: 'Standard Buying')."""
    return frappe.db.get_single_value("Buying Settings", "buying_price_list") or "Standard Buying"


def _default_uom() -> str:
    """UdM de respaldo cuando el ítem no tiene stock_uom (caso excepcional)."""
    return frappe.db.get_single_value("Stock Settings", "stock_uom") or "Unit"


def _get_credit_to(company: str, currency: str = "GTQ") -> str:
    return (
        frappe.db.get_value(
            "Account",
            {"company": company, "account_type": "Payable",
             "account_currency": currency, "is_group": 0},
            "name",
            order_by="creation asc",
        )
        or frappe.db.get_value("Company", company, "default_payable_account")
        or ""
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_purchase_tax_templates(company: str = None) -> list:
    """Plantillas de impuestos de compra realmente activas y ligadas a la compañía conectada."""
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)
    return [
        {
            "name":       t.name,
            "is_default": int(t.is_default or 0),
            "rate":       _get_purchase_tax_rate(t.name),
        }
        for t in _list_company_tax_templates(company)
    ]


@frappe.whitelist()
def get_purchase_defaults(company: str = None) -> dict:
    """Valores por defecto para inicializar el formulario de compra."""
    from facex_multi.api.invoice import get_effective_company
    company  = get_effective_company(company)
    abbr     = _get_abbr(company)
    currency = frappe.db.get_value("Company", company, "default_currency") or "GTQ"

    tax_templates = get_purchase_tax_templates(company)

    return {
        "company":             company,
        "abbr":                abbr,
        "currency":            currency,
        "tax_templates":       tax_templates,
        "default_tax_template": tax_templates[0]["name"] if tax_templates else "",
        "credit_to_gtq":     _get_credit_to(company, "GTQ"),
        "credit_to_usd":     _get_credit_to(company, "USD"),
        "buying_price_list": _get_buying_price_list(),
        "naming_series":     _get_naming_series_pi(company),
        "posting_date":      today(),
        "update_stock":      1,
    }


@frappe.whitelist()
def get_item_purchase_info(item_code: str, company: str = None) -> dict:
    """Info del ítem necesaria para el grid de compras."""
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)

    item = frappe.db.get_value(
        "Item", item_code,
        ["item_name", "item_group", "has_serial_no", "is_stock_item", "stock_uom"],
        as_dict=True,
    )
    if not item:
        frappe.throw(f"Producto '{item_code}' no encontrado.")

    warehouse = _resolve_item_warehouse(item_code, company) if item.is_stock_item else ""

    return {
        "item_code":    item_code,
        "item_name":    item.item_name,
        "item_group":   item.item_group,
        "has_serial_no": int(item.has_serial_no or 0),
        "is_stock_item": int(item.is_stock_item or 0),
        "uom":          item.stock_uom or _default_uom(),
        "warehouse":    warehouse,
    }


@frappe.whitelist()
def search_items(txt: str = "", company: str = None) -> list:
    """Búsqueda de ítems para el grid de compras (por código o nombre)."""
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)

    results = frappe.db.get_all(
        "Item",
        filters=[
            ["disabled", "=", 0],
            ["is_purchase_item", "=", 1],
        ],
        or_filters=[
            ["name", "like", f"%{txt}%"],
            ["item_name", "like", f"%{txt}%"],
        ],
        fields=["name as item_code", "item_name", "item_group",
                "has_serial_no", "is_stock_item", "stock_uom"],
        order_by="item_code asc",
        limit=25,
    )

    out = []
    for r in results:
        wh = _resolve_item_warehouse(r.item_code, company) if r.is_stock_item else ""
        out.append({
            "item_code":    r.item_code,
            "item_name":    r.item_name,
            "item_group":   r.item_group,
            "has_serial_no": int(r.has_serial_no or 0),
            "is_stock_item": int(r.is_stock_item or 0),
            "uom":          r.stock_uom or _default_uom(),
            "warehouse":    wh,
        })
    return out


@frappe.whitelist()
def search_suppliers(txt: str = "", company: str = None) -> list:
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)
    results = frappe.db.sql(
        """
        SELECT name, supplier_name
        FROM `tabSupplier`
        WHERE disabled = 0
          AND supplier_name LIKE %(q)s
          AND (
              bfel_company = %(company)s
              OR (bfel_company IS NULL OR bfel_company = '')
          )
        ORDER BY supplier_name ASC
        LIMIT 20
        """,
        {"q": f"%{txt}%", "company": company},
        as_dict=True,
    )
    return [{"value": r.name, "label": r.supplier_name} for r in results]


@frappe.whitelist()
def search_suppliers_maint(txt: str = "", company: str = None) -> list:
    """Para el módulo de mantenimiento: retorna listado con tax_id."""
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)
    results = frappe.db.sql(
        """
        SELECT name, supplier_name, tax_id
        FROM `tabSupplier`
        WHERE disabled = 0
          AND (name LIKE %(q)s OR supplier_name LIKE %(q)s OR tax_id LIKE %(q)s)
          AND (
              bfel_company = %(company)s
              OR (bfel_company IS NULL OR bfel_company = '')
          )
        ORDER BY supplier_name ASC
        LIMIT 30
        """,
        {"q": f"%{txt}%", "company": company},
        as_dict=True,
    )
    return [{"name": r.name, "supplier_name": r.supplier_name, "tax_id": r.tax_id or ""} for r in results]


@frappe.whitelist()
def get_supplier(name: str, company: str = None) -> dict:
    """Retorna los campos relevantes del proveedor para el formulario de mantenimiento."""
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)
    doc = frappe.get_doc("Supplier", name)
    if doc.get("bfel_company") and doc.bfel_company != company:
        frappe.throw(f"El proveedor '{name}' pertenece a otra compañía.")
    return {
        "name":          doc.name,
        "supplier_name": doc.supplier_name or "",
        "tax_id":        doc.tax_id or "",
        "custom_direccion": doc.get("custom_direccion") or "",
        "custom_telefono":  doc.get("custom_telefono") or "",
    }


@frappe.whitelist()
def create_or_update_supplier(data_json: str, company: str = None) -> dict:
    """Crea o actualiza un proveedor desde el mantenimiento FacEx."""
    from facex_multi.api.invoice import get_effective_company
    data    = json.loads(data_json) if isinstance(data_json, str) else data_json
    name    = (data.get("name") or "").strip()
    company = get_effective_company(company)

    if name:
        doc = frappe.get_doc("Supplier", name)
        if doc.get("bfel_company") and doc.bfel_company != company:
            frappe.throw("No tiene permisos para modificar un proveedor de otra compañía.")
    else:
        doc = frappe.new_doc("Supplier")
        doc.supplier_type  = "Company"
        doc.supplier_group = (
            frappe.db.get_value("Supplier Group", {"is_group": 0}, "name", order_by="lft asc")
            or "All Supplier Groups"
        )
        if doc.meta.has_field("bfel_company"):
            doc.bfel_company = company

    for field in ("supplier_name", "tax_id", "custom_direccion", "custom_telefono"):
        if field in data:
            try:
                setattr(doc, field, data[field])
            except Exception:
                pass

    if doc.meta.has_field("bfel_company"):
        doc.bfel_company = company

    doc.save(ignore_permissions=False)
    frappe.db.commit()
    return {"name": doc.name, "supplier_name": doc.supplier_name}


@frappe.whitelist()
def get_purchase_list(
    company: str = None,
    start_date: str = None,
    end_date: str = None,
    supplier: str = None,
    docstatus: str = None,
    limit: int = 50,
) -> list:
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)

    filters = [["company", "=", company]]
    if start_date and end_date:
        filters.append(["posting_date", "between", [start_date, end_date]])
    elif start_date:
        filters.append(["posting_date", ">=", start_date])
    elif end_date:
        filters.append(["posting_date", "<=", end_date])
    if supplier:
        filters.append(["supplier", "=", supplier])
    if docstatus is not None and docstatus != "":
        filters.append(["docstatus", "=", int(docstatus)])

    return frappe.get_all(
        "Purchase Invoice",
        filters=filters,
        fields=["name", "supplier", "posting_date", "bill_no", "bill_date",
                "grand_total", "outstanding_amount", "docstatus", "currency"],
        order_by="posting_date desc, creation desc",
        limit=int(limit),
    )


@frappe.whitelist()
def get_purchase_invoice(name: str, company: str = None) -> dict:
    from facex_multi.api.invoice import get_effective_company
    company = get_effective_company(company)
    doc = frappe.get_doc("Purchase Invoice", name.strip())
    if doc.company != company and frappe.session.user != "Administrator":
        frappe.throw("No tiene permisos para ver esta factura de compra.")
    d = doc.as_dict()
    for item in d.get("items", []):
        item["has_serial_no"] = int(
            frappe.db.get_value("Item", item.get("item_code"), "has_serial_no") or 0
        )
    return d


@frappe.whitelist()
def save_purchase_invoice(data_json: str) -> dict:
    """
    Crea o actualiza un Purchase Invoice en borrador.
    data_json: {name?, company, supplier, posting_date, bill_no, bill_date,
                currency, tax_type (nombre real de la Plantilla de Impuestos de Compra),
                bfel_multi_tipo (Tipo FEL de encabezado),
                items: [{item_code, qty, rate, warehouse, serial_no, bfel_multi_tipo}]}
    """
    data     = json.loads(data_json) if isinstance(data_json, str) else data_json
    from facex_multi.api.invoice import get_effective_company
    company       = get_effective_company(data.get("company"))
    currency      = data.get("currency") or "GTQ"
    tax_template  = data.get("tax_type") or ""
    name          = (data.get("name") or "").strip()

    if name and frappe.db.exists("Purchase Invoice", name):
        doc = frappe.get_doc("Purchase Invoice", name)
        if doc.docstatus != 0:
            frappe.throw("Solo se pueden editar facturas en borrador.")
        doc.items = []
        doc.taxes = []
    else:
        doc = frappe.new_doc("Purchase Invoice")
        ns  = _get_naming_series_pi(company)
        doc.naming_series = ns[0] if ns else ""

    doc.company           = company
    doc.supplier          = data["supplier"]
    doc.posting_date      = data.get("posting_date") or today()
    doc.bill_no           = data.get("bill_no") or ""
    doc.bill_date         = data.get("bill_date") or doc.posting_date
    doc.currency             = currency
    doc.conversion_rate      = 1
    doc.price_list_currency  = currency
    doc.plc_conversion_rate  = 1
    doc.buying_price_list    = _get_buying_price_list()
    doc.update_stock         = 1
    doc.taxes_and_charges    = _resolve_purchase_tax_template(company, tax_template)
    doc.credit_to            = _get_credit_to(company, currency)
    doc.set_warehouse        = None
    doc.bfel_multi_tipo      = data.get("bfel_multi_tipo") or ""

    for row in data.get("items", []):
        item_code = (row.get("item_code") or "").strip()
        if not item_code:
            continue
        qty       = flt(row.get("qty") or 1)
        rate      = flt(row.get("rate") or 0)
        wh        = row.get("warehouse") or _resolve_item_warehouse(item_code, company)
        serial_no = (row.get("serial_no") or "").strip()

        item_info  = frappe.db.get_value(
            "Item", item_code, ["has_serial_no", "is_stock_item", "stock_uom"], as_dict=True
        )
        has_serial = int(item_info.has_serial_no or 0)
        is_stock   = int(item_info.is_stock_item or 0)
        stock_uom  = item_info.stock_uom or _default_uom()

        if has_serial and serial_no:
            serials   = [s.strip() for s in serial_no.replace(",", "\n").split("\n") if s.strip()]
            qty       = len(serials)
            serial_no = "\n".join(serials)

        item_row = {
            "item_code":         item_code,
            "qty":               qty,
            "rate":              rate,
            "amount":            qty * rate,
            "uom":               stock_uom,
            "stock_uom":         stock_uom,
            "conversion_factor": 1.0,
            "bfel_multi_tipo":   row.get("bfel_multi_tipo") or "",
            "warehouse":         wh if is_stock else "",
        }
        if has_serial and serial_no:
            item_row["serial_no"] = serial_no

        doc.append("items", item_row)

    if not doc.items:
        frappe.throw("Agregue al menos un producto antes de guardar.")

    doc.taxes = []
    for tax_row in _get_tax_rows(doc.taxes_and_charges):
        doc.append("taxes", tax_row)

    doc.flags.ignore_validate   = True
    doc.flags.ignore_mandatory  = True

    # validate() está deshabilitado (ignore_validate) para evitar las validaciones
    # estrictas de ERPNext no aplicables a este flujo simplificado; por eso se
    # invoca manualmente el cálculo de impuestos/totales que normalmente correría
    # dentro de validate(), para que la plantilla de impuestos sí se refleje en
    # total_taxes_and_charges/grand_total.
    doc.calculate_taxes_and_totals()

    if name and frappe.db.exists("Purchase Invoice", name):
        doc.save(ignore_permissions=True)
    else:
        doc.insert(ignore_permissions=True)

    frappe.db.commit()
    return {"success": True, "name": doc.name}


@frappe.whitelist()
def submit_purchase_invoice(name: str) -> dict:
    doc = frappe.get_doc("Purchase Invoice", name.strip())
    if doc.docstatus != 0:
        frappe.throw("La factura ya fue validada o cancelada.")
    doc.flags.ignore_validate = False
    doc.submit()
    frappe.db.commit()
    return {"success": True, "name": doc.name}


@frappe.whitelist()
def cancel_purchase_invoice(name: str) -> dict:
    doc = frappe.get_doc("Purchase Invoice", name.strip())
    if doc.docstatus != 1:
        frappe.throw("Solo se pueden cancelar facturas validadas.")
    doc.cancel()
    frappe.db.commit()
    return {"success": True, "name": doc.name}


@frappe.whitelist()
def process_purchase_excel(file_url: str, company: str = None) -> dict:
    """
    Parsea Excel con 2 hojas:
      Hoja 1 – ENCABEZADO (fila 2): proveedor | fecha_registro | no_factura | fecha_factura | moneda
      Hoja 2 – DETALLE (desde fila 2): codigo_item | precio_unitario | serie
    Una fila por unidad para ítems con serie; varias filas sin serie se acumulan.
    """
    import openpyxl
    from frappe.utils.file_manager import get_file_path
    from facex_multi.api.invoice import get_effective_company

    company   = get_effective_company(company)
    file_path = get_file_path(file_url)
    wb        = openpyxl.load_workbook(file_path, data_only=True)

    # ── Hoja 1: Encabezado ──────────────────────────────────────────
    ws_h   = wb.worksheets[0]
    rows_h = list(ws_h.iter_rows(min_row=2, values_only=True))
    if not rows_h or not rows_h[0][0]:
        frappe.throw("La hoja ENCABEZADO está vacía o mal formada.")
    h = rows_h[0]

    def _val(v, default=""):
        return str(v).strip() if v is not None else default

    header = {
        "supplier":     _val(h[0] if len(h) > 0 else ""),
        "posting_date": _val(h[1], today()),
        "bill_no":      _val(h[2] if len(h) > 2 else ""),
        "bill_date":    _val(h[3] if len(h) > 3 else None, _val(h[1], today())),
        "currency":     _val(h[4] if len(h) > 4 else "GTQ", "GTQ").upper() or "GTQ",
    }

    # ── Hoja 2: Detalle ─────────────────────────────────────────────
    ws_d   = wb.worksheets[1]
    rows_d = list(ws_d.iter_rows(min_row=2, values_only=True))

    # Agrupar por item_code manteniendo orden de aparición
    order  = []
    groups = {}
    for row in rows_d:
        if not row or not row[0]:
            continue
        item_code = str(row[0]).strip()
        rate      = flt(row[1] if len(row) > 1 else 0)
        serial    = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if item_code not in groups:
            order.append(item_code)
            groups[item_code] = {"rate": rate, "serials": [], "qty": 0}
        if serial:
            groups[item_code]["serials"].append(serial)
        else:
            groups[item_code]["qty"] += 1
        if rate:
            groups[item_code]["rate"] = rate

    items = []
    errors = []
    for item_code in order:
        g    = groups[item_code]
        info = frappe.db.get_value(
            "Item", item_code,
            ["item_name", "item_group", "has_serial_no", "is_stock_item"],
            as_dict=True,
        )
        if not info:
            errors.append(f"Producto '{item_code}' no encontrado en el sistema.")
            continue

        has_serial = int(info.has_serial_no or 0)
        wh         = _resolve_item_warehouse(item_code, company) if info.is_stock_item else ""

        if has_serial and g["serials"]:
            qty       = len(g["serials"])
            serial_no = "\n".join(g["serials"])
        else:
            qty       = g["qty"] or 1
            serial_no = ""

        if has_serial and not serial_no:
            errors.append(f"'{item_code}' maneja series pero no se encontraron números de serie.")

        items.append({
            "item_code":     item_code,
            "item_name":     info.item_name,
            "item_group":    info.item_group,
            "has_serial_no": has_serial,
            "qty":           qty,
            "rate":          g["rate"],
            "amount":        qty * g["rate"],
            "warehouse":     wh,
            "serial_no":     serial_no,
        })

    return {"header": header, "items": items, "errors": errors}
